#!/usr/bin/env python3
"""
scripts/xlsx-to-rows.py — BoQ spreadsheet → raw JSON rows (for api/rfq.php).

Generic adapter: detects the header row (contains Description + Qty/Quantity),
maps columns, skips title/banner rows but keeps the current section banner
(all-caps text in the description column) as the row's section tag. Output is
raw-ish rows — api/import.php parse_boq does the normalization + issue flags.

Usage: python3 scripts/xlsx-to-rows.py <file.xlsx> [sheet_name]
"""
import sys, re, json
import openpyxl

path = sys.argv[1]
sheet = sys.argv[2] if len(sys.argv) > 2 else None

wb = openpyxl.load_workbook(path, data_only=True)
sheets = wb.sheetnames
if sheet and sheet in sheets:
    ws = wb[sheet]
else:
    # Pick the sheet with the most data-like rows (skip 'Final Summary' etc.)
    best, best_n = None, -1
    for name in sheets:
        n = 0
        for row in wb[name].iter_rows(values_only=True):
            c0 = '' if row[0] is None else str(row[0]).strip()
            if re.match(r'^\d+(\.\d+)?$', c0):
                n += 1
        if n > best_n:
            best, best_n = name, n
    ws = wb[best]

# ── Find the header row + column map ─────────────────
headers = None
col = {'item': 0, 'desc': 4, 'spec': 5, 'size': 6, 'unit': 7, 'qty': 8}
rows_data = list(ws.iter_rows(values_only=True))
for i, row in enumerate(rows_data[:15]):
    labels = {str(v).strip().lower(): j for j, v in enumerate(row) if v is not None}
    if 'description' in labels or ('desc' in labels and 'qty' in labels):
        headers = i
        col['desc'] = labels.get('description', labels.get('desc', col['desc']))
        col['spec'] = labels.get('spec', labels.get('specification', col['spec']))
        col['size'] = labels.get('size (nb)', labels.get('size', labels.get('nb', col['size'])))
        col['unit'] = labels.get('unit', labels.get('uom', col['unit']))
        col['qty'] = labels.get('qty', labels.get('quantity', labels.get('qty.', col['qty'])))
        col['item'] = labels.get('item', labels.get('item no.', labels.get('item no', col['item'])))
        break

def cell(row, idx):
    if idx >= len(row):
        return ''
    v = row[idx]
    return '' if v is None else str(v).strip()

out = []
section = ''
for i, row in enumerate(rows_data):
    if headers is not None and i <= headers:
        continue
    desc = cell(row, col['desc'])
    if not desc:
        continue
    # Section banner: all-caps text (SECTION 1 / CLARIFIED WATER B4 / BILL NO 1 / NOTES)
    if re.match(r'^[A-Z][A-Z0-9 &\/\-\:]{2,40}$', desc) and len(desc) < 45:
        if not re.match(r'^(item|drawing|p&id|description|spec|qty)', desc.lower()):
            section = desc
            continue
    # Item-number rows: leading number in the item column (or desc starts with a number)
    item = cell(row, col['item'])
    is_data = bool(re.match(r'^\d+(\.\d+)?$', item))
    if not is_data and not re.match(r'^\d+(\.\d+)?\s', desc):
        continue  # noise (blank, notes, images)
    out.append({
        'item_no': item,
        'desc': desc,
        'spec': cell(row, col['spec']),
        'size': cell(row, col['size']),
        'unit': cell(row, col['unit']),
        'qty': cell(row, col['qty']),
        'section': section,
    })

print(json.dumps(out))

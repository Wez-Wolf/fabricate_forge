#!/usr/bin/env python3
"""
xlsx-cells.py — BoQ spreadsheet → raw CELL grid (for the cell-based source picker).

Unlike xlsx-to-rows.py (which hard-codes a column map and emits normalized
row objects), this emits the *raw* cells with A1 coordinates so a UI can show
the spreadsheet and let the user pick which columns/cells become data.

Usage:
  python3 scripts/xlsx-cells.py <file.xlsx> [sheet_name] [max_rows]

Output JSON:
{
  "sheets": ["Sheet A", "Sheet B", ...],
  "active": "Sheet A",                 # the sheet used
  "headerRow": 8,                      # guessed header row (1-based), 0 if none
  "headerLabels": {"A":"ABC3", ...},   # values on the guessed header row
  "cells": [                           # flat list, only non-empty cells
    {"r": 8, "c": 1, "coord": "A8", "v": "ABC3"},
    ...
  ]
}
"""
import sys, json
import openpyxl

A1 = lambda c: openpyxl.utils.get_column_letter(c)

def col_to_idx(letter):
    n = 0
    for ch in letter:
        n = n * 26 + (ord(ch.upper()) - 64)
    return n

def find_active(wb):
    """Prefer the sheet with the most data-like rows (skip 'Final Summary' etc)."""
    best, best_n = None, -1
    for name in wb.sheetnames:
        ws = wb[name]
        n = sum(1 for row in ws.iter_rows(values_only=True)
                if any(v is not None and str(v).strip() != '' for v in row))
        if n > best_n:
            best, best_n = name, n
    return best or (wb.sheetnames[0] if wb.sheetnames else None)

HEADER_HINTS = ('description', 'item', 'qty', 'quantity', 'size', 'spec', 'abc', 'cost type')

def cell_value(v):
    """read_only mode yields <ReadOnlyCell>/<EmptyCell> placeholders for sparse
    regions; reduce those to their underlying value or None."""
    if v is None:
        return None
    s = str(v)
    if 'ReadOnlyCell' in s or 'EmptyCell' in s:
        # e.g. "<ReadOnlyCell 'Sheet'.M1>" — no real value
        return None
    return s.strip() if isinstance(v, str) else (
        v.isoformat() if hasattr(v, 'isoformat') else str(v))

def guess_header_row(ws):
    best, best_score = 0, 0
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=25, values_only=True), 1):
        vals = [str(v).strip().lower() for v in row
                if v is not None and cell_value(v) not in (None, '')]
        joined = ' '.join(vals)
        score = sum(1 for h in HEADER_HINTS if h in joined)
        if score > best_score:
            best, best_score = i, score
    return best

def main():
    if len(sys.argv) < 2:
        sys.stderr.write("usage: xlsx-cells.py <file.xlsx> [sheet] [max_rows]\n")
        sys.exit(2)
    path = sys.argv[1]
    # Remaining args: a non-numeric token is the sheet name; a numeric token
    # is max_rows. Type-based (not positional) so either/both can be omitted.
    sheet_arg = None
    max_rows = 0
    for a in sys.argv[2:]:
        if a.isdigit():
            max_rows = int(a)
        else:
            sheet_arg = a
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = wb.sheetnames
    active = sheet_arg if sheet_arg in sheets else find_active(wb)
    ws = wb[active]

    header_row = guess_header_row(ws)
    cells = []
    for ri, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        if max_rows and ri > max_rows:
            break
        for ci, v in enumerate(row, 1):
            s = cell_value(v)
            if not s:
                continue
            cells.append({"r": ri, "c": ci, "coord": f"{A1(ci)}{ri}", "v": s})

    header_labels = {}
    if header_row:
        for c in cells:
            if c['r'] == header_row:
                header_labels[c['coord'][:-len(str(header_row))]] = c['v']

    json.dump({
        "sheets": sheets,
        "active": active,
        "headerRow": header_row,
        "headerLabels": header_labels,
        "cells": cells,
    }, sys.stdout, ensure_ascii=False)

if __name__ == "__main__":
    main()
